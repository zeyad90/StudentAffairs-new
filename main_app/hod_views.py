import json
import requests
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (HttpResponse, HttpResponseRedirect,
                              get_object_or_404, redirect, render)
from django.templatetags.static import static
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import UpdateView

from .forms import *
from .models import *


def admin_home(request):
    total_staff = Staff.objects.all().count()
    total_students = Student.objects.all().count()
    subjects = Subject.objects.all()
    total_subject = subjects.count()
    total_course = Semester.objects.all().count()
    attendance_list = Attendance.objects.filter(subject__in=subjects)
    total_attendance = attendance_list.count()
    attendance_list = []
    subject_list = []
    for subject in subjects:
        attendance_count = Attendance.objects.filter(subject=subject).count()
        subject_list.append(subject.name[:7])
        attendance_list.append(attendance_count)
    # Total Subjects and students in Each Course
    course_all = Semester.objects.all()
    course_name_list = []
    subject_count_list = []
    student_count_list_in_course = []

    for course in course_all:
        subjects = Subject.objects.filter(Semester_id=course.id).count()
        students = Student.objects.filter(Semester_id=course.id).count()
        course_name_list.append(course.name)
        subject_count_list.append(subjects)
        student_count_list_in_course.append(students)
    subject_all = Subject.objects.all()
    subject_list = []
    student_count_list_in_subject = []
    for subject in subject_all:
        course = Semester.objects.get(id=subject.Semester.id)
        student_count = Student.objects.filter(Semester_id=course.id).count()
        subject_list.append(subject.name)
        student_count_list_in_subject.append(student_count)


    # For Students
    student_attendance_present_list=[]
    student_attendance_leave_list=[]
    student_name_list=[]

    students = Student.objects.all()
    for student in students:
        
        attendance = AttendanceReport.objects.filter(student_id=student.id, status=True).count()
        absent = AttendanceReport.objects.filter(student_id=student.id, status=False).count()
        leave = LeaveReportStudent.objects.filter(student_id=student.id, status=1).count()
        student_attendance_present_list.append(attendance)
        student_attendance_leave_list.append(leave+absent)
        student_name_list.append(student.admin.first_name)

    context = {
        'page_title': "Administrative Dashboard",
        'total_students': total_students,
        'total_staff': total_staff,
        'total_course': total_course,
        'total_subject': total_subject,
        'subject_list': subject_list,
        'attendance_list': attendance_list,
        'student_attendance_present_list': student_attendance_present_list,
        'student_attendance_leave_list': student_attendance_leave_list,
        "student_name_list": student_name_list,
        "student_count_list_in_subject": student_count_list_in_subject,
        "student_count_list_in_course": student_count_list_in_course,
        "course_name_list": course_name_list,

    }
    return render(request, 'hod_template/home_content.html', context)


def add_staff(request):
    form = StaffForm(request.POST or None, request.FILES or None)
    context = {'form': form, 'page_title': 'Add Staff'}
    if request.method == 'POST':
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password')
            course = form.cleaned_data.get('Semester')
            idnumber = form.cleaned_data.get('idnumber')
            try:
                user = CustomUser.objects.create_user(email=email, password=password, user_type=2, first_name=first_name, last_name=last_name,idnumber=idnumber)
        
                user.staff.Semester = course
                user.save()
                messages.success(request, "Successfully Added")
                return redirect(reverse('add_staff'))

            except Exception as e:
                messages.error(request, "Could Not Add " + str(e))
        else:
            messages.error(request, "Please fulfil all requirements")

    return render(request, 'hod_template/add_staff_template.html', context)

def add_student(request):
    student_form = StudentForm(request.POST or None, request.FILES or None)
    context = {'form': student_form, 'page_title': 'Add Student'}
    if request.method == 'POST':
        if student_form.is_valid():
            first_name = student_form.cleaned_data.get('first_name')
            last_name = student_form.cleaned_data.get('last_name')
            email = student_form.cleaned_data.get('email')
            password = student_form.cleaned_data.get('password')
            course = student_form.cleaned_data.get('Semester')
            idnumber = student_form.cleaned_data.get('idnumber')
            try:
                user = CustomUser.objects.create_user(
                    email=email, password=password, user_type=3, first_name=first_name, last_name=last_name,idnumber=idnumber)
                user.student.Semester = course
                user.save()
                messages.success(request, "Successfully Added")
                # Send an email to the new student's email address
                subject = 'Welcome to our school!'
                message = f'Hi {first_name},\n\nWelcome to our school! We look forward to having you as a student in {course}.\n\nBest regards,\nThe School Team. \nyou can login to your account use your NPA email and your password is {password}'
                from_email = 'admin@myschool.com'
                recipient_list = ['zeyadnaji90@gmail.com']
                send_mail(subject, message, from_email, recipient_list)
                return redirect(reverse('add_student'))
            except Exception as e:
                messages.error(request, "Could Not Add: " + str(e))
        else:
            messages.error(request, "Could Not Add: ")
    return render(request, 'hod_template/add_student_template.html', context)

def add_course(request):
    form = CourseForm(request.POST or None)
    context = {'form': form,'page_title': 'Add Semseter'}
    if request.method == 'POST':
        if form.is_valid():
            course_name = form.cleaned_data['name']
            if Semester.objects.filter(name=course_name).exists():
                messages.error(request, "Course with this name already exists")
            else:
                course = form.save(commit=False)
                course.save()
                messages.success(request, "Successfully Added")
                return redirect(reverse('add_course'))
        else:
            messages.error(request, "Could Not Add")
    return render(request, 'hod_template/add_course_template.html', context)


def add_subject(request):
    form = SubjectForm(request.POST or None)
    context = {
        'form': form,
        'page_title': 'Add Course'
    }
    if request.method == 'POST':
        if form.is_valid():
            name = form.cleaned_data.get('name')
            course = form.cleaned_data.get('Semester')
            staff = form.cleaned_data.get('staff')
            try:
                subject = Subject()
                subject.name = name
                subject.staff = staff
                subject.Semester = course
                subject.save()
                messages.success(request, "Successfully Added")
                return redirect(reverse('add_subject'))

            except Exception as e:
                messages.error(request, "Could Not Add " + str(e))
        else:
            messages.error(request, "Fill Form Properly")

    return render(request, 'hod_template/add_subject_template.html', context)


def manage_staff(request):
    allStaff = CustomUser.objects.filter(user_type=2)
    context = {
        'allStaff': allStaff,
        'page_title': 'Manage Staff'
    }
    return render(request, "hod_template/manage_staff.html", context)


def manage_student(request):
    students = CustomUser.objects.filter(user_type=3)
    context = {
        'students': students,
        'page_title': 'Manage Students'
    }
    return render(request, "hod_template/manage_student.html", context)


def manage_course(request):
    courses = Semester.objects.all()
    context = {
        'courses': courses,
        'page_title': 'Manage Semseter'
    }
    return render(request, "hod_template/manage_course.html", context)


def manage_subject(request):
    subjects = Subject.objects.all()
    context = {
        'subjects': subjects,
        'page_title': 'Manage Course'
    }
    return render(request, "hod_template/manage_subject.html", context)





def edit_staff(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    form = StaffForm(request.POST or None, instance=staff)
    context = {
        'form': form,
        'staff_id': staff_id,
        'page_title': 'Edit Staff'
    }
    if request.method == 'POST':
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            username = form.cleaned_data.get('username')
            idnumber = form.cleaned_data.get('idnumber')
            se = form.cleaned_data.get('Semester')
            email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password') or None
            try:
                user = CustomUser.objects.get(id=staff.admin.id)
                user.username = username
                user.email = email
                if password != None:
                    user.set_password(password)
                user.first_name = first_name
                user.last_name = last_name
                staff.Semester=se;
                staff.idnumber=idnumber
                user.save()
                staff.save()
                messages.success(request, "Successfully Updated")
                return redirect(reverse('edit_staff', args=[staff_id]))
            except Exception as e:
                messages.error(request, "Could Not Update " + str(e))
        else:
            messages.error(request, "Please fil form properly")
    else:
        user = CustomUser.objects.get(id=staff.admin.id)
        staff = Staff.objects.get(id=staff.id)
        return render(request, "hod_template/edit_staff_template.html", context)

def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    form = StudentForm(request.POST or None, instance=student)
    context = {
        'form': form,
        'student_id': student_id,
        'page_title': 'Edit Student'
    }
    if request.method == 'POST':
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            username = form.cleaned_data.get('username')
            email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password') or None
            course = form.cleaned_data.get('Semester')
            try:
                user = CustomUser.objects.get(id=student.admin.id)
                user.username = username
                user.email = email
                if password != None:
                    user.set_password(password)
                user.first_name = first_name
                user.last_name = last_name
                student.Semester = course
                user.save()
                student.save()
                messages.success(request, "Successfully Updated")
                return redirect(reverse('edit_student', args=[student_id]))
            except Exception as e:
                messages.error(request, "Could Not Update " + str(e))
        else:
            messages.error(request, "Please Fill Form Properly!")
    else:
        return render(request, "hod_template/edit_student_template.html", context)

def edit_course(request, course_id):
    instance = get_object_or_404(Semester, id=course_id)
    form = CourseForm(request.POST or None, instance=instance)
    context = {
        'form': form,
        'course_id': course_id,
        'page_title': 'Edit Course'
    }
    if request.method == 'POST':
        if form.is_valid():
            name = form.cleaned_data.get('name')
            try:
                course = Semester.objects.get(id=course_id)
                course.name = name
                course.save()
                messages.success(request, "Successfully Updated")
            except:
                messages.error(request, "Could Not Update")
        else:
            messages.error(request, "Could Not Update")

    return render(request, 'hod_template/edit_course_template.html', context)


def edit_subject(request, subject_id):
    instance = get_object_or_404(Subject, id=subject_id)
    form = SubjectForm(request.POST or None, instance=instance)
    context = {
        'form': form,
        'subject_id': subject_id,
        'page_title': 'Edit Subject'
    }
    if request.method == 'POST':
        if form.is_valid():
            name = form.cleaned_data.get('name')
            course = form.cleaned_data.get('Semester')
            staff = form.cleaned_data.get('staff')
            try:
                subject = Subject.objects.get(id=subject_id)
                subject.name = name
                subject.staff = staff
                subject.Semester = course
                subject.save()
                messages.success(request, "Successfully Updated")
                return redirect(reverse('edit_subject', args=[subject_id]))
            except Exception as e:
                messages.error(request, "Could Not Add " + str(e))
        else:
            messages.error(request, "Fill Form Properly")
    return render(request, 'hod_template/edit_subject_template.html', context)

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic.edit import CreateView
from .models import Student, StudentBulkUpload
import csv
from django.views.generic import View

import openpyxl
from django.shortcuts import render

from django.contrib import messages
from django.shortcuts import redirect
from .models import CustomUser, Semester
import openpyxl


from django.core.mail import send_mail

def process_excel(request):
    if request.method == 'POST' and request.FILES['my_file']:
        excel_file = request.FILES['my_file']
        # Check if the file is an xlsx file
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload an xlsx file')
            return redirect('add_student')
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        # Get the active worksheet
        ws = wb.active
        # Iterate over the rows starting from the second row
        for row in ws.iter_rows(min_row=2, values_only=True):
            first_name = row[0]
            last_name = row[1]
            email = row[2]
            password = row[3]
            idnumber = row[4]
            Semester = row[5]
            # Save the student to the database
            try:
                course_obj = Semester.objects.get(name=Semester) # retrieve the Course instance with the given name
                user = CustomUser.objects.create_user(
                    email=email, password=password, user_type=3, first_name=first_name, last_name=last_name,idnumber=idnumber)
                user.student.course = course_obj # assign the Course instance to the course field
                user.save()
                messages.success(request, "Successfully Added")
                # Send an email to the new student's email address
                subject = 'Welcome to our school!'
                message = f'Hi {first_name},\n\nWelcome to our school! We look forward to having you as a student in {Semester}.\n\nBest regards,\nThe School Team. \nyou can login to your account use your NPA email and your password is {password}'
                from_email = 'admin@myschool.com'
                recipient_list = ['zeyadnaji90@gmail.com']
                send_mail(subject, message, from_email, recipient_list)
            except Exception as e:
                messages.error(request, "Could Not Add: " + str(e))
        return redirect('add_student')
    else:
        messages.error(request, "Could Not Add: ")
        return redirect('add_student')


class StudentBulkUploadView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = StudentBulkUpload
    template_name = "hod_template/students_upload.html"
    fields = ["csv_file"]
    success_url = "../student/manage/"
    success_message = "Successfully uploaded students"


    
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class DownloadCSVViewdownloadcsv(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=student_template.xlsx'
        
        # create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # set the column headers
        headers = [
            'firstName',
            'lastName',
            'email',
            'password',
            'idnumber',
            'Semester',
        ]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
        
        # save the workbook to the response
        wb.save(response)
        return response



@csrf_exempt
def check_email_availability(request):
    email = request.POST.get("email")
    try:
        user = CustomUser.objects.filter(email=email).exists()
        if user:
            return HttpResponse(True)
        return HttpResponse(False)
    except Exception as e:
        return HttpResponse(False)

@csrf_exempt
def view_student_leave(request):
    if request.method != 'POST':
        allLeave = LeaveReportStudent.objects.all()
        context = {
            'allLeave': allLeave,
            'page_title': 'Leave Applications From Students'
        }
        return render(request, "hod_template/student_leave_view.html", context)
    else:
        id = request.POST.get('id')
        status = request.POST.get('status')
        if (status == '1'):
            status = 1
        else:
            status = -1
        try:
            leave = get_object_or_404(LeaveReportStudent, id=id)
            leave.status = status
            leave.save()
            return HttpResponse(True)
        except Exception as e:
            return False


def admin_view_attendance(request):
    subjects = Subject.objects.all()
    context = {
        'subjects': subjects,
        'page_title': 'View Attendance'
    }

    return render(request, "hod_template/admin_view_attendance.html", context)


@csrf_exempt
def get_admin_attendance(request):
    subject_id = request.POST.get('subject')
    attendance_date_id = request.POST.get('attendance_date_id')
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        attendance = get_object_or_404(
            Attendance, id=attendance_date_id)
        attendance_reports = AttendanceReport.objects.filter(
            attendance=attendance)
        json_data = []
        for report in attendance_reports:
            data = {
                "status":  str(report.status),
                "name": str(report.student)
            }
            json_data.append(data)
        return JsonResponse(json.dumps(json_data), safe=False)
    except Exception as e:
        return None


def admin_view_profile(request):
    admin = get_object_or_404(Admin, admin=request.user)
    form = AdminForm(request.POST or None, request.FILES or None,
                     instance=admin)
    context = {'form': form,
               'page_title': 'View/Edit Profile'
               }
    if request.method == 'POST':
        try:
            if form.is_valid():
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                password = form.cleaned_data.get('password') or None
                custom_user = admin.admin
                if password != None:
                    custom_user.set_password(password)
                custom_user.first_name = first_name
                custom_user.last_name = last_name
                custom_user.save()
                messages.success(request, "Profile Updated!")
                return redirect(reverse('admin_view_profile'))
            else:
                messages.error(request, "Invalid Data Provided")
        except Exception as e:
            messages.error(
                request, "Error Occured While Updating Profile " + str(e))
    return render(request, "hod_template/admin_view_profile.html", context)


def delete_staff(request, staff_id):
    staff = get_object_or_404(CustomUser, staff__id=staff_id)
    staff.delete()
    messages.success(request, "Staff deleted successfully!")
    return redirect(reverse('manage_staff'))


def delete_student(request, student_id):
    student = get_object_or_404(CustomUser, student__id=student_id)
    student.delete()
    messages.success(request, "Student deleted successfully!")
    return redirect(reverse('manage_student'))


def delete_course(request, course_id):
    course = get_object_or_404(Semester, id=course_id)
    try:
        course.delete()
        messages.success(request, "Course deleted successfully!")
    except Exception:
        messages.error(
            request, "Sorry, some students are assigned to this course already. Kindly change the affected student course and try again")
    return redirect(reverse('manage_course'))


def delete_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    subject.delete()
    messages.success(request, "Subject deleted successfully!")
    return redirect(reverse('manage_subject'))

